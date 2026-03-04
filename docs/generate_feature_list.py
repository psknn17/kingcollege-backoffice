"""
Generate Feature List Excel for Payment Backoffice System
Matching the format of the example file: Feature lists_02032026.xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── COLOR CONSTANTS ──────────────────────────────────────────────────────────
COLOR_TITLE_BG   = "FF1F2329"   # dark charcoal (title row)
COLOR_TITLE_FG   = "FFFFFFFF"   # white
COLOR_HEADER_BG  = "FFE1EAFF"   # light blue (column headers)
COLOR_MODULE_BG  = "FFDEE0E3"   # light gray (module group rows)
COLOR_DEFAULT_FG = "FF000000"   # black

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=COLOR_DEFAULT_FG, size=9.75, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(wrap=True, vertical="top", horizontal="left"):
    return Alignment(wrap_text=wrap, vertical=vertical, horizontal=horizontal)

def thin_border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── DATA DEFINITION ──────────────────────────────────────────────────────────
# (wbs, feature_text, description, resource_link)
# wbs = single letter  → MODULE ROW (gray bg, bold)
# wbs = X#             → SUB-MODULE ROW (indent 1)
# wbs = X#.#           → FEATURE ROW (indent 2)
# wbs = X#.#.#         → DETAIL ROW (indent 3)

DATA = [
    # ── A: LOGIN ──────────────────────────────────────────────────────────────
    ("A", "เข้าสู่ระบบ (Login)", "", ""),
    ("A1", "เข้าสู่ระบบ", "", ""),
    ("A1.1",
        "ผู้ใช้เข้าสู่ระบบด้วย Username และ Password",
        "รองรับ 4 บทบาท:\n1. Admin\n2. Admin Accountant\n3. Viewer\n4. Approver",
        ""),
    ("A1.2",
        "Redirect ตามบทบาทอัตโนมัติ",
        "- Admin / Admin Accountant → Dashboard\n- Viewer → Dashboard (read-only)\n- Approver → Approval Queue",
        ""),
    ("A1.3",
        "ป้องกันการเข้าถึงหน้าที่ไม่มีสิทธิ์",
        "Approver เข้าได้เฉพาะ Approval Queue, User Profile, Settings, Activity\nหน้าอื่นถูก redirect กลับ",
        ""),

    # ── B: DASHBOARD ──────────────────────────────────────────────────────────
    ("B", "Dashboard & รายงาน (Dashboard & Reports)", "", ""),
    ("B1", "Dashboard ภาพรวมรายได้", "", ""),
    ("B1.1",
        "แสดงยอดรายได้รวมแยกตามโมดูล",
        "โมดูลที่แสดง: Tuition, ECA, Trip, Exam, Bus, External\nรูปแบบ: Bar chart, Pie chart, Summary cards",
        ""),
    ("B1.2",
        "กรองข้อมูลตามปีการศึกษาและเทอม",
        "- เลือกปีการศึกษา\n- เลือกเทอม (Term 1 / 2 / 3)\n- ปุ่ม Reset filter",
        ""),
    ("B1.3",
        "แสดงสถิติวิธีการชำระเงิน",
        "Credit Card, PromptPay, Bank Counter, WeChat Pay, Alipay",
        ""),
    ("B2", "รายงานส่วนลด (Discount Reports)", "", ""),
    ("B2.1",
        "แสดงการใช้งานส่วนลดตามกลุ่มนักเรียน",
        "วิเคราะห์ผลกระทบรายได้จากส่วนลดแต่ละประเภท",
        ""),
    ("B2.2",
        "กรองข้อมูลตามปีการศึกษาและเทอม",
        "- เลือกปีการศึกษา\n- เลือกเทอม\n- ปุ่ม Reset filter",
        ""),

    # ── C: PAYMENT REMINDERS ──────────────────────────────────────────────────
    ("C", "แจ้งเตือนชำระเงิน (Payment Reminders)", "", ""),
    ("C1", "ตั้งค่าแจ้งเตือนค้างชำระ (Debt Reminder Settings)", "", ""),
    ("C1.1",
        "สร้าง / แก้ไข / ลบ การตั้งค่าแจ้งเตือน",
        "มี 4 template สำเร็จรูป:\n1. First Reminder\n2. Second Reminder\n3. Final Notice\n4. Overdue Notice",
        ""),
    ("C1.2",
        "กำหนด Subject, Title และ Message ของอีเมล",
        "รองรับ Placeholder: {parent_name}, {student_name}, {amount}, {due_date}, {days_remaining}",
        ""),
    ("C1.3",
        "เลือกช่องทางส่ง: Email / SMS / ทั้งคู่",
        "", ""),
    ("C1.4",
        "กำหนดวันที่และเวลาส่ง (Scheduled Send)",
        "รองรับ: ส่งทันที หรือ กำหนดเวลาล่วงหน้า",
        ""),
    ("C1.5",
        "ติดตามสถานะ: Draft / Scheduled / Sent / Cancelled",
        "", ""),
    ("C2", "แม่แบบใบแจ้งหนี้ / ใบเสร็จ (Invoice/Receipt Template)", "", ""),
    ("C2.1",
        "แก้ไข HTML template สำหรับอีเมลใบแจ้งหนี้และใบเสร็จ",
        "รองรับ Tab แยก Edit / Preview\nดู Live Preview ก่อนบันทึก",
        ""),
    ("C2.2",
        "บันทึกและ Reset template กลับค่าเริ่มต้น",
        "", ""),
    ("C3", "ประวัติการส่งอีเมล (Email History)", "", ""),
    ("C3.1",
        "ดูประวัติการส่งอีเมลทั้งหมด",
        "แสดง: ผู้รับ, ประเภท, วัน-เวลา, สถานะ (Sent/Failed/Bounced)",
        ""),
    ("C3.2",
        "ค้นหาตามผู้รับ / ประเภท / หัวข้ออีเมล",
        "", ""),
    ("C3.3",
        "Export ประวัติการส่งเป็นไฟล์ Excel",
        "", ""),
    ("C4", "ประวัติการชำระเงิน (Payment History)", "", ""),
    ("C4.1",
        "ดูรายการชำระเงินทั้งหมดในระบบ",
        "แสดง: วันที่, นักเรียน, ยอดเงิน, วิธีชำระ, สถานะ",
        ""),
    ("C4.2",
        "กรองตามวันที่ / สถานะ / นักเรียน / ยอดเงิน",
        "", ""),
    ("C4.3",
        "จัดเรียงข้อมูลตามคอลัมน์ และแบ่งหน้า",
        "", ""),
    ("C4.4",
        "Export ข้อมูลเป็นไฟล์ Excel",
        "", ""),

    # ── D: TUITION ────────────────────────────────────────────────────────────
    ("D", "ค่าเทอม (Tuition Management)", "", ""),
    ("D1", "ค่าเทอมตามปีการศึกษา (Tuition By Year)", "", ""),
    ("D1.1",
        "แก้ไขอัตราค่าเทอมแยกตามชั้นเรียนและปีการศึกษา",
        "รองรับ 16 ชั้น: Pre-Nursery ถึง Year 13\nแบ่ง 3 เทอม: Term 1, 2, 3",
        ""),
    ("D1.2",
        "บันทึกการเปลี่ยนแปลงทั้งหมดพร้อมกัน",
        "กด Save all เพื่อบันทึกทุกชั้นพร้อมกัน",
        ""),
    ("D1.3",
        "เลือกแสดงข้อมูลตามปีการศึกษา",
        "", ""),
    ("D2", "กลุ่มส่วนลดนักเรียน (Student Discount Groups)", "", ""),
    ("D2.1",
        "เพิ่ม / แก้ไข / ลบ กลุ่มส่วนลด",
        "กำหนดส่วนลดแบบ: เปอร์เซ็นต์ (%) หรือ จำนวนเงินคงที่ (Fixed Amount)",
        ""),
    ("D2.2",
        "เปิด / ปิด การใช้งานกลุ่มส่วนลด",
        "", ""),
    ("D3", "จัดการใบแจ้งหนี้ค่าเทอม (Invoice Management - Tuition)", "", ""),
    ("D3.1",
        "ดู / เพิ่ม / แก้ไข / ลบ ใบแจ้งหนี้",
        "สถานะใบแจ้งหนี้: Draft → Pending → Approved / Rejected → Sent → Paid / Overdue / Cancelled",
        ""),
    ("D3.2",
        "ค้นหาตามเลขที่ / ชื่อนักเรียน / สถานะ / วันที่",
        "", ""),
    ("D3.3",
        "จัดเรียงข้อมูลตามคอลัมน์ และแบ่งหน้า",
        "", ""),
    ("D3.4",
        "ดำเนินการหลายรายการพร้อมกัน (Bulk Actions)",
        "รองรับ: อนุมัติ, ปฏิเสธ, ทำเครื่องหมายชำระแล้ว, ส่งอีเมล, Export PDF",
        ""),
    ("D3.5",
        "นำเข้าข้อมูลจาก Interface File (Excel)",
        "รองรับรูปแบบ Interface File สำหรับเชื่อมต่อระบบบัญชี",
        ""),
    ("D3.6",
        "Export ใบแจ้งหนี้เป็น Excel / ZIP / PDF",
        "", ""),
    ("D3.7",
        "ส่งใบแจ้งหนี้ทางอีเมล พร้อมบันทึก Log",
        "", ""),
    ("D3.8",
        "สร้าง PDF ใบแจ้งหนี้",
        "รวม: ข้อมูลโรงเรียน, รายการ, ส่วนลด, บัญชีธนาคาร",
        ""),
    ("D4", "รายการและแม่แบบ (Items & Templates)", "", ""),
    ("D4.1",
        "เพิ่ม / แก้ไข / ลบ รายการค่าใช้จ่าย",
        "ฟิลด์: รหัส, ชื่อ, คำอธิบาย, ยอดเงิน, หมวดหมู่, Nominal Code",
        ""),
    ("D4.2",
        "นำเข้าข้อมูลจาก Excel / CSV พร้อม Preview และตรวจสอบ Duplicate",
        "แสดง Warning box สีเหลืองเมื่อพบรายการซ้ำ",
        ""),
    ("D4.3",
        "Export รายการทั้งหมดเป็น Excel",
        "", ""),
    ("D4.4",
        "ค้นหาตามรหัส / ชื่อ / ชั้นเรียน และแบ่งหน้า",
        "", ""),
    ("D4.5",
        "Sync ราคากับข้อมูล Tuition By Year",
        "อัพเดทราคาเท่านั้น ไม่สร้างรายการใหม่อัตโนมัติ",
        ""),
    ("D4.6",
        "เปิด / ปิด การใช้งานรายการ (Active/Inactive)",
        "", ""),
    ("D5", "ใบเสร็จค่าเทอม (Receipts - Tuition)", "", ""),
    ("D5.1",
        "ดูและบริหารจัดการใบเสร็จรับเงิน",
        "สถานะ: Issued, Resent, Failed",
        ""),
    ("D5.2",
        "กรองตามสถานะ / ชั้น / เลขที่ใบเสร็จ / วันที่",
        "", ""),
    ("D5.3",
        "ดำเนินการหลายรายการพร้อมกัน (Bulk Actions)",
        "รองรับ: อนุมัติ, แก้ไข, ลบ",
        ""),
    ("D5.4",
        "สร้าง PDF ใบเสร็จรับเงิน",
        "", ""),
    ("D5.5",
        "นำเข้าและ Export ข้อมูลใบเสร็จ",
        "", ""),
    ("D6", "ใบลดหนี้ (Credit Notes)", "", ""),
    ("D6.1",
        "เพิ่ม / แก้ไข / ลบ ใบลดหนี้",
        "ประเภท: Refund, Adjustment, Cancellation\nสถานะ: Draft, Issued, Applied, Cancelled",
        ""),
    ("D6.2",
        "ค้นหาตามเลขที่ / ชื่อนักเรียน / สถานะ และแบ่งหน้า",
        "", ""),
    ("D6.3",
        "นำเข้าและ Export ข้อมูลใบลดหนี้เป็น Excel",
        "", ""),

    # ── E: ECA ────────────────────────────────────────────────────────────────
    ("E", "กิจกรรมนอกเวลา (ECA / After School)", "", ""),
    ("E1", "ใบแจ้งหนี้ ECA", "", ""),
    ("E1.1", "CRUD, Bulk Actions, Export, Email, PDF — เหมือน Tuition Invoice", "หมวดหมู่: ECA (เช่น ดนตรี, กีฬา, ชมรม)", ""),
    ("E2", "รายการและแม่แบบ ECA", "", ""),
    ("E2.1", "CRUD, Import, Export — เหมือน Tuition Items", "มีตัวอย่างรายการ: Piano, Guitar, Violin ฯลฯ", ""),
    ("E3", "ใบเสร็จ ECA", "", ""),
    ("E3.1", "ดู, Bulk, Export, PDF — เหมือน Tuition Receipt", "", ""),
    ("E4", "กลุ่มส่วนลด ECA", "", ""),
    ("E4.1", "CRUD กลุ่มส่วนลดเฉพาะหมวด ECA", "", ""),

    # ── F: TRIP ───────────────────────────────────────────────────────────────
    ("F", "ทริปและกิจกรรม (Trip & Activity)", "", ""),
    ("F1", "ใบแจ้งหนี้ทริป", "", ""),
    ("F1.1", "CRUD, Bulk Actions, Export, Email, PDF — เหมือน Tuition Invoice", "หมวดหมู่: trip", ""),
    ("F2", "รายการและแม่แบบทริป", "", ""),
    ("F2.1", "CRUD, Import, Export — เหมือน Tuition Items", "", ""),
    ("F3", "ใบเสร็จทริป", "", ""),
    ("F3.1", "ดู, Bulk, Export, PDF — เหมือน Tuition Receipt", "", ""),
    ("F4", "กลุ่มส่วนลดทริป", "", ""),
    ("F4.1", "CRUD กลุ่มส่วนลดเฉพาะหมวด Trip", "", ""),

    # ── G: EXAM ───────────────────────────────────────────────────────────────
    ("G", "ค่าสอบ (Exam)", "", ""),
    ("G1", "ใบแจ้งหนี้ค่าสอบ", "", ""),
    ("G1.1", "CRUD, Bulk Actions, Export, Email, PDF — เหมือน Tuition Invoice", "หมวดหมู่: exam", ""),
    ("G2", "รายการและแม่แบบค่าสอบ", "", ""),
    ("G2.1", "CRUD, Import, Export — เหมือน Tuition Items", "", ""),
    ("G3", "ใบเสร็จค่าสอบ", "", ""),
    ("G3.1", "ดู, Bulk, Export, PDF — เหมือน Tuition Receipt", "", ""),
    ("G4", "กลุ่มส่วนลดค่าสอบ", "", ""),
    ("G4.1", "CRUD กลุ่มส่วนลดเฉพาะหมวด Exam", "", ""),

    # ── H: BUS ────────────────────────────────────────────────────────────────
    ("H", "รถรับส่ง (School Bus)", "", ""),
    ("H1", "ใบแจ้งหนี้รถรับส่ง", "", ""),
    ("H1.1", "CRUD, Bulk Actions, Export, Email, PDF — เหมือน Tuition Invoice", "หมวดหมู่: bus", ""),
    ("H2", "รายการและแม่แบบรถรับส่ง", "", ""),
    ("H2.1", "CRUD, Import, Export — เหมือน Tuition Items", "รายการตามเส้นทาง/จุดรับส่ง", ""),
    ("H3", "ใบเสร็จรถรับส่ง", "", ""),
    ("H3.1", "ดู, Bulk, Export, PDF — เหมือน Tuition Receipt", "", ""),
    ("H4", "กลุ่มส่วนลดรถรับส่ง", "", ""),
    ("H4.1", "CRUD กลุ่มส่วนลดเฉพาะหมวด Bus", "", ""),

    # ── I: EXTERNAL ───────────────────────────────────────────────────────────
    ("I", "ใบแจ้งหนี้บุคคลภายนอก (External Invoice)", "", ""),
    ("I1", "ใบแจ้งหนี้ภายนอก", "", ""),
    ("I1.1",
        "CRUD, Bulk Actions, Export, Email, PDF — เหมือน Tuition Invoice",
        "หมวดหมู่: external\nผู้รับ: บุคคล/องค์กรภายนอก (ไม่ใช่นักเรียน)",
        ""),
    ("I1.2",
        "สร้างใบแจ้งหนี้สำหรับลูกค้าภายนอก",
        "ระบุ: ชื่อผู้รับ, ที่อยู่, อีเมล, ชื่องาน",
        ""),
    ("I2", "รายชื่อลูกค้า (Client List)", "", ""),
    ("I2.1",
        "เพิ่ม / แก้ไข / ลบ ข้อมูลลูกค้า",
        "ฟิลด์: ชื่อบริษัท/บุคคล, ชื่อผู้ติดต่อ, ที่อยู่",
        ""),
    ("I2.2",
        "ค้นหาลูกค้า",
        "", ""),
    ("I2.3",
        "นำเข้าลูกค้าจาก Excel / CSV",
        "Column ที่ต้องการ: Client Name (required), Contact Name, Address\nรองรับ Preview และตรวจสอบ Duplicate",
        ""),
    ("I2.4",
        "Export รายชื่อลูกค้าเป็น Excel",
        "Export ตามรายการที่ filter อยู่ปัจจุบัน",
        ""),
    ("I3", "รายการและแม่แบบภายนอก", "", ""),
    ("I3.1", "CRUD, Import, Export — เหมือน Tuition Items", "ประเภท: Rental, Catering, Service, Event, Other", ""),
    ("I4", "ใบเสร็จภายนอก", "", ""),
    ("I4.1", "ดู, Bulk, Export, PDF — เหมือน Tuition Receipt", "", ""),
    ("I5", "กลุ่มส่วนลดภายนอก", "", ""),
    ("I5.1", "CRUD กลุ่มส่วนลดเฉพาะหมวด External", "", ""),

    # ── J: STUDENT MANAGEMENT ─────────────────────────────────────────────────
    ("J", "ข้อมูลนักเรียน (Student Management)", "", ""),
    ("J1", "รายชื่อนักเรียน (Student List)", "", ""),
    ("J1.1",
        "เพิ่ม / แก้ไข / ลบ ข้อมูลนักเรียน",
        "ฟิลด์: ชื่อ, รหัส, ชั้น, วันเกิด, เพศ, สถานะ, ปีการศึกษา, ครอบครัว, ผู้ปกครอง",
        ""),
    ("J1.2",
        "ค้นหาตามชื่อ / รหัส และกรองตามสถานะ / ชั้น / ปีการศึกษา",
        "", ""),
    ("J1.3",
        "จัดเรียงข้อมูลตามคอลัมน์ และแบ่งหน้า",
        "", ""),
    ("J1.4",
        "นำเข้านักเรียนจาก Excel / CSV",
        "Column ที่ต้องการ: Student ID, First Name (หรือ Name), Year Group, Academic Year (หรือ Year)\nแสดง Preview 10 แถวแรก + Toggle แสดงทั้งหมด",
        ""),
    ("J1.5",
        "Export รายชื่อนักเรียนเป็น Excel (25 columns)",
        "รวม: ข้อมูลส่วนตัว, ครอบครัว, ผู้ปกครอง 2 คน, วันลงทะเบียน",
        ""),
    ("J2", "กลุ่มครอบครัว (Family Groups)", "", ""),
    ("J2.1",
        "เพิ่ม / แก้ไข / ลบ ข้อมูลครอบครัว",
        "รองรับหลายบุตรต่อครอบครัว\nฟิลด์: Family Code, ชื่อครอบครัว, ที่อยู่, Email สำหรับใบแจ้งหนี้",
        ""),
    ("J2.2",
        "ค้นหาและกรองข้อมูลครอบครัว และแบ่งหน้า",
        "", ""),
    ("J2.3",
        "Export ข้อมูลครอบครัวเป็น Excel",
        "", ""),

    # ── K: DISCOUNT ───────────────────────────────────────────────────────────
    ("K", "การจัดการส่วนลด (Discount Management)", "", ""),
    ("K1", "โปรโมชั่น (Promotional Campaigns)", "", ""),
    ("K1.1",
        "เพิ่ม / แก้ไข / ลบ โปรโมชั่น",
        "กำหนด: ชื่อ, ประเภทส่วนลด (% หรือ Fixed), วันเริ่ม-สิ้นสุด, สถานะ",
        ""),
    ("K1.2",
        "ติดตามสถานะ: Active / Ended / Scheduled",
        "", ""),
    ("K2", "จัดการยกเว้นค่าเทอม (Waive Fee Management)", "", ""),
    ("K2.1",
        "ดูรายชื่อนักเรียนที่มีสิทธิ์ยกเว้นค่าเทอม",
        "ประเภท: Sibling Discount, Staff Child, Scholarship, Early Bird",
        ""),
    ("K2.2",
        "กรองตามสถานะและปีการศึกษา",
        "", ""),
    ("K3", "รายละเอียดยกเว้นรายปี (Waive Fee Year Details)", "", ""),
    ("K3.1",
        "เพิ่ม / แก้ไข รายการยกเว้นค่าเทอมรายปี",
        "ค้นหาตามชื่อนักเรียน / สถานะ",
        ""),

    # ── L: APPROVAL ───────────────────────────────────────────────────────────
    ("L", "การอนุมัติและสร้างใบแจ้งหนี้ (Approval & Invoice Creation)", "", ""),
    ("L1", "คิวอนุมัติ (Approval Queue)", "", ""),
    ("L1.1",
        "ดูรายการใบแจ้งหนี้ที่รออนุมัติ",
        "แสดงสถานะ: Wait / Approved / Rejected\nเลขที่ใบแจ้งหนี้แสดงเฉพาะ Approved",
        ""),
    ("L1.2",
        "อนุมัติหรือปฏิเสธใบแจ้งหนี้ พร้อมระบุเหตุผล",
        "บันทึก: Approved by, Approved at, Rejected reason",
        ""),
    ("L1.3",
        "กรองตามสถานะ / วันที่ / ชื่อนักเรียน และแบ่งหน้า",
        "", ""),
    ("L1.4",
        "ดำเนินการหลายรายการพร้อมกัน (Bulk Approve/Reject)",
        "", ""),
    ("L1.5",
        "Export ใบแจ้งหนี้เป็น Excel / PDF",
        "", ""),
    ("L1.6",
        "ส่งอีเมลแจ้งผลการอนุมัติ",
        "", ""),
    ("L2", "สร้างใบแจ้งหนี้นักเรียน (Invoice Creation)", "", ""),
    ("L2.1",
        "ค้นหาและเลือกนักเรียน",
        "", ""),
    ("L2.2",
        "เลือกรายการค่าใช้จ่าย",
        "เลือกจากรายการที่มีอยู่ หรือสร้างรายการใหม่",
        ""),
    ("L2.3",
        "คำนวณส่วนลดอัตโนมัติ 7 ประเภท",
        "1. Student Group\n2. Sibling\n3. Staff Child\n4. Scholarship\n5. Early Bird\n6. Fixed Amount\n7. Percentage",
        ""),
    ("L2.4",
        "กำหนดวันครบกำหนดชำระ และหมายเหตุ",
        "", ""),
    ("L2.5",
        "บันทึก Draft และ Submit เพื่ออนุมัติ",
        "", ""),
    ("L2.6",
        "สร้างและดาวน์โหลด PDF ใบแจ้งหนี้",
        "", ""),
    ("L2.7",
        "Export Interface File สำหรับระบบบัญชี",
        "", ""),
    ("L2.8",
        "บันทึกและเรียกใช้ Template ใบแจ้งหนี้",
        "", ""),
    ("L3", "สร้างใบแจ้งหนี้ภายนอก (External Invoice Creation)", "", ""),
    ("L3.1",
        "ระบุข้อมูลผู้รับ (ชื่อ, ที่อยู่, อีเมล)",
        "", ""),
    ("L3.2",
        "เลือกรายการ, คำนวณยอด, สร้าง PDF",
        "", ""),

    # ── M: SETTINGS ───────────────────────────────────────────────────────────
    ("M", "ตั้งค่าระบบ (Settings & Administration)", "", ""),
    ("M1", "ตั้งค่าโรงเรียน (School Settings)", "", ""),
    ("M1.1",
        "แก้ไขข้อมูลโรงเรียน",
        "ฟิลด์: ชื่อ (EN/TH), ที่อยู่, เบอร์โทร, อีเมล, เว็บไซต์, Tax ID, โลโก้",
        ""),
    ("M1.2",
        "ตั้งค่าบัญชีธนาคารสำหรับใบแจ้งหนี้",
        "ฟิลด์: ธนาคาร, ชื่อบัญชี, เลขบัญชี, สาขา, SWIFT Code",
        ""),
    ("M1.3",
        "Reset ข้อมูลระบบ (เลือกเฉพาะบางส่วน)",
        "คงไว้: users, activityLogs, authUser, app-language\nลบ: ใบแจ้งหนี้, ค่าเทอม, ส่วนลด",
        ""),
    ("M2", "ตั้งค่าบัญชีธนาคาร (Bank Settings)", "", ""),
    ("M2.1",
        "เพิ่ม / แก้ไข / ลบ บัญชีธนาคาร",
        "รองรับหลายบัญชี กำหนดบัญชีหลัก (Primary)",
        ""),
    ("M3", "ตั้งค่าเทอมเรียน (Term Settings)", "", ""),
    ("M3.1",
        "เพิ่ม / แก้ไข ปีการศึกษาและเทอม",
        "กำหนด: วันเริ่ม-สิ้นสุดแต่ละเทอม\nตรวจสอบ: วันที่ทับซ้อน, ปีต่อเนื่อง, สูงสุด 3 เทอม/ปี",
        ""),
    ("M4", "จัดการผู้ใช้งาน (User Management)", "", ""),
    ("M4.1",
        "เพิ่ม / แก้ไข / ลบ ผู้ใช้งาน",
        "ฟิลด์: Username, Email, ชื่อ-นามสกุล, Role, สถานะ",
        ""),
    ("M4.2",
        "กำหนด Role: Admin / Admin Accountant / Viewer / Approver",
        "", ""),
    ("M4.3",
        "ค้นหาตามชื่อ / Role / อีเมล และแบ่งหน้า",
        "", ""),
    ("M4.4",
        "ส่งอีเมล Reset รหัสผ่าน",
        "", ""),
    ("M5", "บันทึกกิจกรรมระบบ (Activity Log)", "", ""),
    ("M5.1",
        "ดูประวัติการดำเนินการทั้งหมด",
        "บันทึก: ผู้ใช้, Action, Module, วัน-เวลา",
        ""),
    ("M5.2",
        "ค้นหาและกรองประวัติ และแบ่งหน้า",
        "", ""),
    ("M6", "โปรไฟล์และการตั้งค่าผู้ใช้", "", ""),
    ("M6.1",
        "แก้ไขข้อมูลส่วนตัว และเปลี่ยนรหัสผ่าน",
        "", ""),
    ("M6.2",
        "เลือกภาษา: ไทย / อังกฤษ",
        "", ""),
    ("M6.3",
        "ดูประวัติกิจกรรมส่วนตัว",
        "", ""),
]

# ─── BUILD EXCEL ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "✅ Feature List"

# Column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 65
ws.column_dimensions["D"].width = 30

# ── METADATA ROWS ─────────────────────────────────────────────────────────────
meta = [
    ("Feature List — Payment Backoffice System", None, None, None),
    ("Document updated:", None, datetime(2026, 3, 2), None),
    ("Project Name:", None, "Payment Backoffice System v1.0", None),
    ("Prepared by:", None, "Development Team", None),
    ("Role:",
     "1. Admin\n2. Admin Accountant\n3. Viewer\n4. Approver",
     None, None),
    ("Module:",
     "A. Login\nB. Dashboard & Reports\nC. Payment Reminders\nD. Tuition\nE. ECA\nF. Trip\nG. Exam\nH. Bus\nI. External Invoice\nJ. Student Management\nK. Discount Management\nL. Approval & Invoice Creation\nM. Settings",
     None, None),
]

for i, row_data in enumerate(meta, start=1):
    ws.append(row_data)
    row = ws[i]
    if i == 1:
        # Title row
        ws.row_dimensions[i].height = 22
        for cell in row[:4]:
            cell.fill = fill(COLOR_TITLE_BG)
            cell.font = font(bold=True, color=COLOR_TITLE_FG, size=13)
            cell.alignment = align(wrap=False, vertical="center")
        # Merge A1:D1
        ws.merge_cells(f"A1:D1")
    elif i in (5, 6):
        ws.row_dimensions[i].height = 70 if i == 6 else 45
        for cell in row[:2]:
            cell.font = font(bold=(cell.column == 1), size=9.75)
            cell.alignment = align(wrap=True, vertical="top")
    else:
        ws.row_dimensions[i].height = 15
        for cell in row[:4]:
            cell.font = font(size=9.75)
            cell.alignment = align(wrap=False, vertical="center")

# ── COLUMN HEADER ROW ─────────────────────────────────────────────────────────
ws.append(("WBS", "Feature", "Description", "หมายเหตุ / Resource Link"))
header_row = ws[7]
ws.row_dimensions[7].height = 18
for cell in header_row:
    cell.fill = fill(COLOR_HEADER_BG)
    cell.font = font(bold=True, size=9.75)
    cell.alignment = align(wrap=False, vertical="center", horizontal="center")
    cell.border = thin_border()

# ── DATA ROWS ─────────────────────────────────────────────────────────────────
def is_module(wbs):
    return len(wbs) == 1 and wbs.isalpha()

def is_submodule(wbs):
    # e.g. A1, B2, D3 — letter + single digit (no dot)
    return len(wbs) >= 2 and wbs[0].isalpha() and "." not in wbs

def indent_level(wbs):
    if is_module(wbs): return 0
    dots = wbs.count(".")
    if dots == 0: return 1   # A1, B2
    if dots == 1: return 2   # A1.1, B2.3
    if dots == 2: return 3   # E1.2.1
    return 4

INDENT = "    "  # 4 spaces per level

for wbs, feature, description, resource in DATA:
    lvl = indent_level(wbs)
    indented_feature = (INDENT * lvl) + feature if lvl > 0 else feature

    ws.append((wbs, indented_feature, description, resource))
    cur_row = ws.max_row

    if is_module(wbs):
        ws.row_dimensions[cur_row].height = 18
        for col in range(1, 5):
            cell = ws.cell(row=cur_row, column=col)
            cell.fill = fill(COLOR_MODULE_BG)
            cell.font = font(bold=True, size=9.75)
            cell.alignment = align(wrap=False, vertical="center")
            cell.border = thin_border()
    else:
        # Auto-height: count newlines
        lines = max(
            feature.count("\n") + 1,
            description.count("\n") + 1,
            1
        )
        ws.row_dimensions[cur_row].height = max(15, lines * 15)
        for col in range(1, 5):
            cell = ws.cell(row=cur_row, column=col)
            cell.font = font(size=9.75)
            cell.alignment = align(wrap=True, vertical="top")
            cell.border = thin_border()

# ── FREEZE PANES ──────────────────────────────────────────────────────────────
ws.freeze_panes = "A8"

# ── SAVE ──────────────────────────────────────────────────────────────────────
output_path = "/Users/passkornnabangchang/Desktop/warp/kingcollenge 2/Kingcollegebackoffice-main/docs/Feature_List_PaymentBackoffice_02032026.xlsx"
wb.save(output_path)
print(f"✅ Saved: {output_path}")
