"""Generate Analytics Dashboard business logic spec as Excel."""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY    = "1F3864"   # dark header bg
BLUE    = "2F5496"   # section header bg
LBLUE   = "D6E4F0"  # alternating row
WHITE   = "FFFFFF"
YELLOW  = "FFF2CC"   # highlight
ORANGE  = "F4B942"   # accent border
BLACK   = "000000"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style="thin", color="BFBFBF")
thick = Side(style="medium", color=ORANGE)

def border_all_thin():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def border_header():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def write_header(ws, row, col, text, span=1, bg=NAVY, fg=WHITE, size=12):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=size)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border_header()
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )
    return cell

def write_section(ws, row, col, text, span=1):
    write_header(ws, row, col, text, span=span, bg=BLUE, fg=WHITE, size=11)

def write_cell(ws, row, col, text, bold=False, bg=WHITE, wrap=True, h="left"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(bg)
    cell.font = font(bold=bold)
    cell.alignment = align(h, "center", wrap=wrap)
    cell.border = border_all_thin()
    return cell

def write_row(ws, row, values, bgs=None, bold=False):
    for i, val in enumerate(values):
        bg = (bgs[i] if bgs else WHITE)
        write_cell(ws, row, i + 1, val, bold=bold, bg=bg)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Overview
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Overview"

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 55
ws1.row_dimensions[1].height = 36

write_header(ws1, 1, 1, "Analytics Dashboard — Business Logic Specification", span=2, size=14)

r = 3
write_section(ws1, r, 1, "Filter Controls", span=2); r += 1
write_cell(ws1, r, 1, "Year Filter", bold=True, bg=LBLUE)
write_cell(ws1, r, 2, "All Years / ปีการศึกษา เช่น 2025/2026, 2024/2025"); r += 1
write_cell(ws1, r, 1, "Term Filter", bold=True, bg=LBLUE)
write_cell(ws1, r, 2, "All Terms / Term 1 / Term 2 / Term 3"); r += 1
write_cell(ws1, r, 1, "Fee Type Filter", bold=True, bg=LBLUE)
write_cell(ws1, r, 2, "All Fee Types / Tuition / ECA / Trip / Exam / Bus"); r += 1

r += 1
write_section(ws1, r, 1, "Stat Cards (แถวบนสุด)", span=2); r += 1
stats = [
    ("Gross Revenue", "รายได้รวมก่อนหักส่วนลด"),
    ("Net Revenue",   "รายได้สุทธิหลังหักส่วนลดและ Bank Fees"),
    ("Bank Fees",     "ค่าธรรมเนียมธนาคาร (Online Payment เท่านั้น: Thai QR, Online Credit Card)"),
    ("Students",      "จำนวนนักเรียนทั้งหมด (unique) ที่มี Invoice ในช่วงที่ filter"),
    ("Transactions",  "จำนวน Transaction ทั้งหมด"),
    ("Success Rate",  "อัตรา Transaction สำเร็จ (%)"),
]
for i, (k, v) in enumerate(stats):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws1, r, 1, k, bold=True, bg=bg)
    write_cell(ws1, r, 2, v, bg=bg)
    r += 1

r += 1
write_section(ws1, r, 1, "Invoice ที่นับเป็นรายได้", span=2); r += 1
write_cell(ws1, r, 1, "Status ที่นับ", bold=True, bg=LBLUE)
write_cell(ws1, r, 2, "paid, sent, overdue, approved"); r += 1
write_cell(ws1, r, 1, "Status ที่ไม่นับ", bold=True, bg=LBLUE)
write_cell(ws1, r, 2, "draft, cancelled, rejected, wait"); r += 1
write_cell(ws1, r, 1, "ปีปัจจุบัน (Current Year)", bold=True, bg=LBLUE)
write_cell(ws1, r, 2, "= ปีการศึกษาล่าสุดที่มีข้อมูลใน Invoice (ตรวจจากข้อมูลจริงในระบบ)"); r += 1

r += 1
write_section(ws1, r, 1, "Navigation Tabs", span=2); r += 1
tabs = [
    ("Tab 1", "Revenue Comparison — เปรียบเทียบรายได้ YoY / ToT"),
    ("Tab 2", "AVG Amount — รายได้เฉลี่ยต่อนักเรียน"),
    ("Tab 3", "No. of Transactions — จำนวน Transaction แยก Payment Method"),
    ("Tab 4", "Declined vs Successful — สถานะ Transaction"),
    ("Tab 5", "Bank Fees — ค่าธรรมเนียมธนาคาร"),
    ("Tab 6", "Net vs Gross Revenue — Waterfall ส่วนลด"),
]
for i, (k, v) in enumerate(tabs):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws1, r, 1, k, bold=True, bg=bg)
    write_cell(ws1, r, 2, v, bg=bg)
    r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Filter Logic
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Filter Logic")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 40
ws2.row_dimensions[1].height = 36

write_header(ws2, 1, 1, "เงื่อนไข Filter — ทุก Tab", span=4, size=13)

r = 3
write_section(ws2, r, 1, "กฎหลัก (General Rule)", span=4); r += 1
write_header(ws2, r, 1, "Year Filter", bg=BLUE, fg=WHITE)
write_header(ws2, r, 2, "Term Filter", bg=BLUE, fg=WHITE)
write_header(ws2, r, 3, "ผลลัพธ์", bg=BLUE, fg=WHITE, span=2); r += 1

rules = [
    ("All Years", "All Terms",  "แสดงข้อมูลปีการศึกษาปัจจุบัน (ปีล่าสุดในระบบ)"),
    ("Filter Year", "All Terms", "แสดงปีที่เลือก + ย้อนหลัง 2 ปี (รวม 3 ปี)"),
    ("All Years", "Filter Term", "แสดง Term นั้น + ย้อนหลัง 2 ปี (รวม 3 ปี)"),
    ("Filter Year", "Filter Term", "แสดง Term นั้นใน 3 ปี โดยใช้ปีที่เลือกเป็นฐาน"),
]
for i, (y, t, res) in enumerate(rules):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws2, r, 1, y, bold=True, bg=bg)
    write_cell(ws2, r, 2, t, bold=True, bg=bg)
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    write_cell(ws2, r, 3, res, bg=bg)
    r += 1

r += 1
write_section(ws2, r, 1, "Tab 1 — Revenue Comparison", span=4); r += 1
write_header(ws2, r, 1, "Filter", bg=BLUE, fg=WHITE)
write_header(ws2, r, 2, "Table/Chart", bg=BLUE, fg=WHITE)
write_header(ws2, r, 3, "Columns ที่แสดง", bg=BLUE, fg=WHITE, span=2); r += 1

t1_rules = [
    ("All Years + All Terms", "Bar Chart",                   "Revenue ปีปัจจุบัน (1 set of bars per Year Group)"),
    ("Filter Year",           "Bar Chart",                   "Grouped bars 3 ปี เปรียบเทียบ (3 bars per Year Group)"),
    ("Filter Term",           "Bar Chart",                   "Revenue ปีปัจจุบัน Term นั้น (1 bar per Year Group)"),
    ("All Years + All Terms", "Compare by Term",             "Terms ของปีปัจจุบัน เช่น 2025/26 T1, T2, T3"),
    ("Filter Year",           "Compare by Term",             "ทุก Term × 3 ปี เช่น 23/24 T1, 24/25 T1, 25/26 T1, T2, T3"),
    ("Filter Term",           "Compare by Term",             "Term นั้นใน 3 ปี เช่น 23/24 T1, 24/25 T1, 25/26 T1"),
    ("All Years + All Terms", "Compare by Academic Year",    "ปีปัจจุบัน (1 column)"),
    ("Filter Year",           "Compare by Academic Year",    "3 ปี: ปีที่เลือก + 2 ปีก่อน"),
    ("Filter Term",           "Compare by Academic Year",    "ไม่แสดงข้อมูล (—)"),
]
for i, (f, tbl, cols) in enumerate(t1_rules):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws2, r, 1, f, bold=True, bg=bg)
    write_cell(ws2, r, 2, tbl, bg=bg)
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    write_cell(ws2, r, 3, cols, bg=bg)
    r += 1

r += 1
write_section(ws2, r, 1, "Tab 2 — AVG Amount", span=4); r += 1
write_header(ws2, r, 1, "Filter", bg=BLUE, fg=WHITE)
write_header(ws2, r, 2, "Table", bg=BLUE, fg=WHITE)
write_header(ws2, r, 3, "Columns ที่แสดง", bg=BLUE, fg=WHITE, span=2); r += 1

t2_rules = [
    ("All Years + All Terms", "Avg Revenue by Term",          "Terms ของปีปัจจุบัน"),
    ("Filter Year",           "Avg Revenue by Term",          "ทุก Term × 3 ปี"),
    ("Filter Term",           "Avg Revenue by Term",          "Term นั้นใน 3 ปี"),
    ("All Years + All Terms", "Avg Revenue by Academic Year", "ปีปัจจุบัน (1 column)"),
    ("Filter Year",           "Avg Revenue by Academic Year", "3 ปี"),
    ("Filter Term",           "Avg Revenue by Academic Year", "Term นั้น × 3 ปี"),
]
for i, (f, tbl, cols) in enumerate(t2_rules):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws2, r, 1, f, bold=True, bg=bg)
    write_cell(ws2, r, 2, tbl, bg=bg)
    ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    write_cell(ws2, r, 3, cols, bg=bg)
    r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — AVG Amount Detail
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("AVG Amount Detail")
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 40
ws3.row_dimensions[1].height = 36

write_header(ws3, 1, 1, "Tab 2: AVG Amount — รายละเอียดการคำนวณ", span=3, size=13)

r = 3
write_section(ws3, r, 1, "โครงสร้างตาราง (ทั้ง 2 ตาราง)", span=3); r += 1
write_header(ws3, r, 1, "Column", bg=BLUE, fg=WHITE)
write_header(ws3, r, 2, "ชื่อ", bg=BLUE, fg=WHITE)
write_header(ws3, r, 3, "คำอธิบาย", bg=BLUE, fg=WHITE); r += 1
cols_detail = [
    ("Fixed", "Year Group", "ชื่อกลุ่มชั้นปี เช่น Pre-Nursery, Year 1 ... Year 13"),
    ("Per Term/Year", "Students", "จำนวนนักเรียน (unique) ที่มี Invoice ใน Term/Year นั้น"),
    ("Per Term/Year", "Total amount", "รายได้รวมทั้งหมดของ Year Group นั้นใน Period นั้น"),
    ("Per Term/Year", "Avg amount", "Total amount ÷ Students = รายได้เฉลี่ยต่อนักเรียน 1 คน"),
]
for i, (col_type, name, desc) in enumerate(cols_detail):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws3, r, 1, col_type, bold=True, bg=bg)
    write_cell(ws3, r, 2, name, bg=bg)
    write_cell(ws3, r, 3, desc, bg=bg)
    r += 1

r += 1
write_section(ws3, r, 1, "แถว Total (ต่อคน)", span=3); r += 1
totals = [
    ("Students", "ผลรวม unique students ใน Term/Year นั้น (ทุก Year Group)"),
    ("Total amount", "ผลรวม Total amount ทุก Year Group"),
    ("Avg amount", "Total amount รวม ÷ Students รวม"),
]
for i, (name, desc) in enumerate(totals):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws3, r, 1, name, bold=True, bg=bg)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    write_cell(ws3, r, 2, desc, bg=bg)
    r += 1

r += 1
write_section(ws3, r, 1, "ตัวอย่างการคำนวณ", span=3); r += 1
write_cell(ws3, r, 1, "Year 7 — 2025/2026 Term 1", bold=True, bg=YELLOW)
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
example = [
    ("Students", "15 คน", ""),
    ("Total amount", "฿825,000", "รายได้รวมทุก Invoice ของ Year 7 ใน Term 1 ปี 2025/26"),
    ("Avg amount", "฿55,000", "825,000 ÷ 15 = 55,000 บาทต่อคน"),
]
for i, (name, val, note) in enumerate(example):
    bg = LBLUE if i % 2 == 0 else WHITE
    write_cell(ws3, r, 1, name, bold=True, bg=bg)
    write_cell(ws3, r, 2, val, bg=bg, h="right")
    write_cell(ws3, r, 3, note, bg=bg)
    r += 1

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Other Tabs
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Other Tabs")
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 60
ws4.row_dimensions[1].height = 36

write_header(ws4, 1, 1, "Tab 3–6: รายละเอียด", span=2, size=13)

r = 3
tabs_detail = [
    ("Tab 3 — No. of Transactions", [
        ("ข้อมูลที่แสดง", "จำนวน Transaction แยกตาม Payment Method"),
        ("ตาราง 1", "Year Group × Payment Method Matrix — นับจำนวน Transaction"),
        ("ตาราง 2", "Payment Method × Academic Year — เปรียบเทียบ YoY"),
        ("Payment Methods", "Bank Transfer, Onsite Credit Card, Thai QR, Online Credit Card, Cheque, Bill Payment"),
    ]),
    ("Tab 4 — Declined vs Successful", [
        ("หมายเหตุ", "ระบบนี้เป็น Offline Payment ทั้งหมด → ไม่มี Declined"),
        ("ข้อมูลที่แสดง", "จำนวน Successful Transaction แยกตาม Payment Method"),
        ("Success Rate", "= 100% เสมอ (ไม่มี Gateway Decline)"),
    ]),
    ("Tab 5 — Bank Fees", [
        ("Online Payment ที่มีค่าธรรมเนียม", "Thai QR, Online Credit Card เท่านั้น"),
        ("ไม่นับ", "Bank Transfer, Cheque, Bill Payment, Onsite Credit Card"),
        ("ตาราง 1", "Bank × Term Matrix — ค่าธรรมเนียมแยกตาม Term"),
        ("ตาราง 2", "Bank × Academic Year — เปรียบเทียบ YoY"),
    ]),
    ("Tab 6 — Net vs Gross Revenue", [
        ("Gross Revenue", "รายได้ก่อนหักส่วนลดทุกประเภท (= Subtotal)"),
        ("Deductions", "ส่วนลด: Sibling Discount, Staff Child, Scholarship, Student Group, Early Bird"),
        ("Net Revenue", "= Gross Revenue − ผลรวม Deductions"),
        ("ตาราง", "Year Group × Academic Year แสดงทั้ง Gross และ Net เปรียบเทียบ YoY"),
    ]),
]

for tab_name, rows_data in tabs_detail:
    write_section(ws4, r, 1, tab_name, span=2); r += 1
    for i, (k, v) in enumerate(rows_data):
        bg = LBLUE if i % 2 == 0 else WHITE
        write_cell(ws4, r, 1, k, bold=True, bg=bg)
        write_cell(ws4, r, 2, v, bg=bg)
        r += 1
    r += 1

# ── Save ───────────────────────────────────────────────────────────────────────
out = "/Users/passkornnabangchang/Desktop/warp/kingcollenge 2/Kingcollegebackoffice-main/docs/Analytics_Dashboard_Spec.xlsx"
wb.save(out)
print(f"Saved: {out}")
